from datetime import datetime, timedelta
import json
import urllib.request
import time
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

xlsx_header = ['id', '省', '日期', '确诊病例', '死亡病例', '治愈病例']

geo_dict = [{'id': 1156420000, 'name': '湖北省'},
            {'id': 1156440000, 'name': '广东省'},
            {'id': 1156410000, 'name': '河南省'},
            {'id': 1156330000, 'name': '浙江省'},
            {'id': 1156430000, 'name': '湖南省'},
            {'id': 1344810000, 'name': '中国香港'},
            {'id': 1156340000, 'name': '安徽省'},
            {'id': 1156360000, 'name': '江西省'},
            {'id': 1156370000, 'name': '山东省'},
            {'id': 1156230000, 'name': '黑龙江省'},
            {'id': 1156320000, 'name': '江苏省'},
            {'id': 1156310000, 'name': '上海市'},
            {'id': 1156110000, 'name': '北京市'},
            {'id': 1156500000, 'name': '重庆市'},
            {'id': 1156510000, 'name': '四川省'},
            {'id': 1158710000, 'name': '中国台湾'},
            {'id': 1156350000, 'name': '福建省'},
            {'id': 1156130000, 'name': '河北省'},
            {'id': 1156610000, 'name': '陕西省'},
            {'id': 1156450000, 'name': '广西壮族自治区'},
            {'id': 1156150000, 'name': '内蒙古自治区'},
            {'id': 1156530000, 'name': '云南省'},
            {'id': 1156120000, 'name': '天津市'},
            {'id': 1156140000, 'name': '山西省'},
            {'id': 1156460000, 'name': '海南省'},
            {'id': 1156520000, 'name': '贵州省'},
            {'id': 1156210000, 'name': '辽宁省'},
            {'id': 1156620000, 'name': '甘肃省'},
            {'id': 1156220000, 'name': '吉林省'},
            {'id': 1156650000, 'name': '新疆维吾尔自治区'},
            {'id': 1156640000, 'name': '宁夏回族自治区'},
            {'id': 1446820000, 'name': '中国澳门'},
            {'id': 1156630000, 'name': '青海省'},
            {'id': 1156540000, 'name': '西藏自治区'}]


def make_xlsx(filename: str, sheet_name: str) -> (Workbook, Worksheet):
    if os.path.exists(filename):
        book = openpyxl.load_workbook(filename)
    else:
        book = openpyxl.Workbook()
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    sheet = book.create_sheet(sheet_name)
    return book, sheet


if __name__ == '__main__':
    full_url = 'https://view.inews.qq.com/g2/getOnsInfo?name=disease_h5&callback=&_=%d' % int(time.time() * 1000)
    current_hour_human = datetime.today().strftime('%Y%m%d-%H')
    current_day_human = datetime.today().strftime('%Y%m%d')
    json_file_path = "/data/covid-19-county-" + current_hour_human + ".json"
    province_stats_file_path = "/data/covid-19-province-" + current_hour_human + ".txt"
    province_xlsx_file_path = "/data/covid-19-province.xlsx"
    # save to file every hour
    response = urllib.request.urlopen(url=full_url)
    result = json.loads(response.read())
    with open(json_file_path, 'w') as json_file:
        json.dump(result, json_file)
    provinces = json.loads(result['data'])['areaTree'][0]['children']
    yesterday_yyyymmdd = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
    unsorted_stats = list()
    # generate table data
    for province in provinces:
        total_stat = province['total']
        province_name = province['name']
        province_id = 0
        for geo in geo_dict:
            if province_name in geo['name']:
                province_id = geo['id']
                province_name = geo['name']
                break
        row = (province_id, province_name, yesterday_yyyymmdd, total_stat['confirm'], total_stat['dead'],
               total_stat['heal'])
        unsorted_stats.append(row)
    ## sort and save to excel(xlsx)
    sorted_stats = sorted(unsorted_stats, key=lambda stats: stats[3], reverse=True)
    xlsx_book_sheet = make_xlsx(province_xlsx_file_path, current_day_human)
    xlsx_book = xlsx_book_sheet[0]
    xlsx_sheet = xlsx_book_sheet[1]
    xlsx_sheet.append(xlsx_header)
    with open(province_stats_file_path, 'w') as province_file:
        province_file.write('\t'.join(map(str,xlsx_header))+'\n')
        for r in sorted_stats:
            xlsx_sheet.append(r)
            printed_row='\t'.join(map(str, r))
            province_file.write(printed_row+'\n')
            print(printed_row)
    xlsx_book.save(province_xlsx_file_path)
    xlsx_book.close()
